import os
from datetime import date, datetime
# Hello manan
import mysql.connector
# Hello from Sehran
from hijri_converter import Hijri, Gregorian
from num2words import num2words
from PyQt5.QtCore import Qt
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtGui import QPixmap
import sys
from PyQt5.QtWidgets import (QApplication,
                             QCheckBox,
                             QComboBox,
                             QDateEdit,
                             QDateTimeEdit,
                             QDial,
                             QDoubleSpinBox,
                             QFontComboBox,
                             QLabel,
                             QLCDNumber,
                             QLineEdit,
                             QMainWindow,
                             QProgressBar,
                             QPushButton,
                             QRadioButton,
                             QSlider,
                             QSpinBox,
                             QTimeEdit,
                             QVBoxLayout,
                             QWidget,
                             QHBoxLayout,
                             QGridLayout,
                             QTableWidget,
                             QHeaderView, QMenu, QAction, QInputDialog, QTableWidgetItem
                             )
from PyQt5.QtGui import (QPalette,
                         QColor,
                         QFont,
                         )

import sys
from random import choice
from googletrans import Translator
from openpyxl import Workbook
import openpyxl

class Settings(QWidget):
    def __init__(self):
        super().__init__()
        
        self.workbook = openpyxl.load_workbook('settings.xlsx')
        self.worksheet = self.workbook.active

        layout = QVBoxLayout()
        self.label = QLabel("Settings")
        self.row1 = QHBoxLayout()
        self.language = QLabel("Language")
        self.chooseLang = QPushButton("Switch to Arabic")

        self.chooseLang.clicked.connect(self.changeLanguage)
        
        self.row1.addWidget(self.language)
        self.row1.addWidget(self.chooseLang)
        
        self.ok = QPushButton("OK")   
        self.ok.clicked.connect(self.closePage)
        self.saveBtn = QPushButton("Save")   
        self.saveBtn.clicked.connect(self.save)

        if(self.worksheet["B2"].value == "arabic"):
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            self.chooseLang.setText("Switch to English/التبديل إلى اللغة الإنجليزية")
            self.saveBtn.setText(ws["B7"].value)
            self.ok.setText(ws["B67"].value)
            self.label.setText(ws["B65"].value)
            self.language.setText(ws["B66"].value)
        else:
            self.chooseLang.setText("Switch to Arabic/التحول إلى اللغة العربية")
        

        layout.addWidget(self.label)
        layout.addLayout(self.row1)
        layout.addWidget(self.saveBtn)
        layout.addWidget(self.ok)
        self.setLayout(layout)

    def closePage(self):
        self.close()
        
    def save(self):
        self.workbook.save('settings.xlsx')
        dialouge = popup()
        dialouge.show()

    def changeLanguage(self, str):
        
        if(self.chooseLang.text() == "Switch to Arabic/التحول إلى اللغة العربية"):
            self.chooseLang.setText("Switch to English/التبديل إلى اللغة الإنجليزية")
            self.worksheet["B2"] = "arabic"
            
        else:
            self.chooseLang.setText("Switch to Arabic/التحول إلى اللغة العربية")
            self.worksheet["B2"] = "english"
            
class popup(QWidget):
    def __init__(self):
        super().__init__()

        layout = QVBoxLayout()
        self.label = QLabel("Restart the app to see changes")
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if(worksheet["B2"].value == "arabic"):
            self.label.setText("أعد تشغيل التطبيق لرؤية التغييرات")
        

        self.okBtn = QPushButton("Ok")

        self.okBtn.clicked.connect(lambda: self.wrapper())

        layout.addWidget(self.label)
        layout.addWidget(self.okBtn)
        self.setLayout(layout)

    def wrapper(self):
        self.close()
