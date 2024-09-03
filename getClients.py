import os
from datetime import date, datetime

import mysql.connector

from hijri_converter import Hijri, Gregorian
from PyQt5.QtCore import Qt
from PyQt5 import QtWidgets
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (QApplication,
                             QCheckBox,
                             QComboBox,
                             QDateEdit,
                             QDateTimeEdit,
                             QDial,
                             QDoubleSpinBox,
                             QFontComboBox,
                             QLabel,
                             QLCDNumber,
                             QLineEdit,
                             QMainWindow,
                             QProgressBar,
                             QPushButton,
                             QRadioButton,
                             QSlider,
                             QSpinBox,
                             QTimeEdit,
                             QVBoxLayout,
                             QWidget,
                             QHBoxLayout,
                             QGridLayout,
                             QTableWidget,
                             QHeaderView, QMenu, QAction, QInputDialog, QFormLayout, QDialogButtonBox, QDialog
                             )
from PyQt5.QtGui import (QPalette,
                         QColor,
                         QFont,
                         )

import sys
from random import choice

from utils import errPopup




class Box(QWidget):

    def __init__(self, color):
        super(Box, self).__init__()
        self.setAutoFillBackground(True)

        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(color))
        self.setPalette(palette)


class MyTable(QTableWidget):
    def __init__(self, r, c):
        super().__init__(r, c)
        self.init_ui()

    def init_ui(self):
        self.show()


class getClients(QMainWindow):

    def __init__(self):
        global mydb
        global mycursor
        import utils
        mydb = utils.connection
        mycursor = mydb.cursor()
        super().__init__()

        self.setWindowTitle("Customers")

        self.setGeometry(150, 150, 800, 300)

        header = QHBoxLayout()
        self.header_label = QLabel("Customers")
        self.header_label.setFixedSize(300, 30)
        self.header_label.setFont(QFont('Arial', 20))

        self.header_label.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        header.addWidget(self.header_label)

        iconBox = QGridLayout()
        iconBox_wid1 = Box("red")
        button_row1 = QHBoxLayout(iconBox_wid1)

        self.accNoCombo = QComboBox()
        self.nameCombo = QComboBox()

        mycursor.execute("SELECT * FROM Customer")
        result = mycursor.fetchall()
        for row in result:
            self.accNoCombo.addItem(str(row[0]))
            self.nameCombo.addItem(row[1])

        self.accNoCombo.currentIndexChanged.connect(lambda: self.updatePage())
        self.nameCombo.currentIndexChanged.connect(lambda: self.updatePage())

        button_row1.addWidget(self.accNoCombo)
        button_row1.addWidget(self.nameCombo)



        iconBox_wid1.setFixedHeight(50)
        iconBox.addWidget(iconBox_wid1, 0, 0)

        query = "SELECT * FROM Customer WHERE accNo = '" + self.accNoCombo.currentText() + "'"
        mycursor.execute(query)
        customer = mycursor.fetchall()

        query = "SELECT SUM(cashDebit), SUM(cashCredit) FROM Cust_Bal WHERE accNo = '" + self.accNoCombo.currentText() + "'"
        mycursor.execute(query)
        customerBal = mycursor.fetchall()

        detBox = QHBoxLayout()


        
        self.phnT = QLabel("Phone No")
        self.emailT = QLabel("Email")
        self.companyT = QLabel("Company")
        self.vatNoT = QLabel("GST No")
        self.addressT = QLabel("Address")
        self.cashDBalT = QLabel("Cash Debit")
        self.cashCBalT = QLabel("Cash Credit")
        # self.cashT = QLabel("Cash Change")
        # self.goldT = QLabel("Gold Change")
        


        form1 = QFormLayout()

        self.phn = QLineEdit()
        self.email = QLineEdit()
        self.company = QLineEdit()
        self.vatNo = QLineEdit()
        self.address = QLineEdit()

        self.cashD = QLabel(str(round(customerBal[0][0], 2)))
        self.cashC = QLabel(str(round(customerBal[0][1], 2)))

        self.phn.setText(str(customer[0][2]))
        self.email.setText(str(customer[0][3]))
        self.company.setText(str(customer[0][4]))
        self.vatNo.setText(str(customer[0][5]))
        self.address.setText(str(customer[0][6]))


        form1.addRow(self.phnT, self.phn)
        form1.addRow(self.emailT, self.email)
        form1.addRow(self.companyT, self.company)
        form1.addRow(self.vatNoT, self.vatNo)
        form1.addRow(self.addressT, self.address)
        form1.addRow(self.cashDBalT, self.cashD)
        form1.addRow(self.cashCBalT, self.cashC)
        # form1.addRow(self.cashBalT, self.cashBal)
        # form1.addRow(self.goldBalT, self.goldBal)

        detBox.addLayout(form1)

        saveBox = QHBoxLayout()
        self.backBtn = QPushButton("Back")
        self.backBtn.clicked.connect(self.close)
        self.saveBtn = QPushButton("Save")
        self.saveBtn.clicked.connect(self.edit)
        saveBox.addWidget(self.backBtn)
        saveBox.addWidget(self.saveBtn)


        mainLayout = QVBoxLayout()
        mainLayout.addLayout(header)
        mainLayout.addLayout(iconBox)
        mainLayout.addLayout(detBox)
        mainLayout.addLayout(saveBox)
        widget = QWidget()
        widget.setLayout(mainLayout)
        self.setCentralWidget(widget)
        self.toggle_translate()

    def toggle_translate(self):
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
            
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            
            self.header_label.setText(ws["B60"].value)
            
            self.phnT.setText(ws["B23"].value)
            self.emailT.setText(ws["B59"].value)
            self.companyT.setText(ws["B72"].value)
            self.vatNoT.setText(ws["B64"].value)
            self.addressT.setText(ws["B74"].value)
            # self.cashT.setText(ws["B75"].value)
            # self.goldT.setText(ws["B76"].value)
            # self.cashBalT.setText(ws["B50"].value)
            # self.goldBalT.setText(ws["B51"].value)
            
            self.saveBtn.setText(ws["B7"].value)
            self.backBtn.setText(ws["B73"].value)

    def updatePage(self):
        print("Working")
        widget = self.sender()
        if widget == self.accNoCombo:
            self.nameCombo.setCurrentIndex(self.accNoCombo.currentIndex())
        if widget == self.nameCombo:
            self.accNoCombo.setCurrentIndex(self.nameCombo.currentIndex())

        query = "SELECT * FROM Customer WHERE accNo = '" + self.accNoCombo.currentText() + "'"
        mycursor.execute(query)
        customer = mycursor.fetchall()

        query = "SELECT SUM(cashDebit), SUM(cashCredit) FROM Cust_Bal WHERE accNo = '" + self.accNoCombo.currentText() + "'"
        mycursor.execute(query)
        customerBal = mycursor.fetchall()

        self.cashD.setText(str(round(customerBal[0][0], 2)))
        self.cashC.setText(str(round(customerBal[0][1],3)))

        self.phn.setText(str(customer[0][2]))
        self.email.setText(str(customer[0][3]))
        self.company.setText(str(customer[0][4]))
        self.vatNo.setText(str(customer[0][5]))
        self.address.setText(str(customer[0][6]))
        # self.cash.setText(str(0))
        # self.gold.setText(str(0))
        self.toggle_translate()

    #
    # add the update function
    #
    
    # edit lock function
    def edit(self):
        # Create a popup using the error popup and change the required
        self.popup = errPopup("1")
        
        self.popup.layout.removeWidget(self.popup.button)
        self.popup.button.deleteLater()
        self.popup.button = None
        
        self.popup.label.setText("Enter the password to change")
        self.okBtn = QPushButton("OK")
        self.cancelBtn = QPushButton("Cancel")
        self.cancelBtn.clicked.connect(self.popup.close)
        password = QLineEdit()
        
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
            
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            
            self.popup.label.setText("أدخل كلمة المرور للتغيير")
            self.okBtn.setText(ws["B67"].value)
            self.cancelBtn.setText(ws["B73"].value)
        
        def verifyi():
            print(password.text())
            if password.text() == "13/12":
                self.popup.close()
                
                self.updateDet()
                
       
        
        self.okBtn.clicked.connect(lambda: verifyi())
        
        self.popup.layout.addWidget(password)
        self.popup.layout.addWidget(self.okBtn)
        self.popup.layout.addWidget(self.cancelBtn)
        self.popup.show()
    
    def updateDet(self):
        
        query = "UPDATE Customer SET phnNo = '" + self.phn.text() + "', email = '" + self.email.text() + "', company = '" + self.company.text() + "', address = '" + self.address.text() + "', GSTIN = '" + self.vatNo.text() + "' WHERE accNo = '" + self.accNoCombo.currentText() + "'"
        mycursor.execute(query)
        mydb.commit()
        # if (int(self.cash.text()) > 0):
        #     query = "INSERT INTO Cust_Bal (date, accNo, cashDebit,goldDebit,cashCredit,goldCredit,billType,billNo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
        #     values = []
        #     values.append(date.today())
        #     values.append(self.accNoCombo.currentText())
        #     values.append(int(self.cash.text()))
        #     values.append(0)
        #     values.append(0)
        #     values.append(0)
        #     values.append("Update cash debit")
        #     values.append("NA")
        #     mycursor.execute(query,values)
        #     mydb.commit()
        # if (int(self.cash.text()) < 0):
        #     query = "INSERT INTO Cust_Bal (date, accNo, cashDebit,goldDebit,cashCredit,goldCredit,billType,billNo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
        #     values = []
        #     values.append(date.today())
        #     values.append(self.accNoCombo.currentText())
            
        #     values.append(0)
        #     values.append(0)
        #     values.append(int(self.cash.text()) * (-1))
        #     values.append(0)
        #     values.append("Update cash credit")
        #     values.append("NA")
        #     mycursor.execute(query,values)
        #     mydb.commit()
        
        # if (int(self.gold.text()) > 0):
        #     query = "INSERT INTO Cust_Bal (date, accNo, cashDebit,goldDebit,cashCredit,goldCredit,billType,billNo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
        #     values = []
        #     values.append(date.today())
        #     values.append(self.accNoCombo.currentText())
        #     values.append(0)
        #     values.append(self.gold.text())
        #     values.append(0)
        #     values.append(0)
        #     values.append("Update gold debit")
        #     values.append("NA")
        #     mycursor.execute(query,values)
        #     mydb.commit()
        # if (int(self.gold.text()) < 0):
        #     query = "INSERT INTO Cust_Bal (date, accNo, cashDebit,goldDebit,cashCredit,goldCredit,billType,billNo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
        #     values = []
        #     values.append(date.today())
        #     values.append(self.accNoCombo.currentText())
            
        #     values.append(0)
        #     values.append(0)
        #     values.append(0)
        #     values.append(int(self.gold.text()) * (-1))
        #     values.append("Update gold credit")
        #     values.append("NA")
        #     mycursor.execute(query,values)
        #     mydb.commit()
        
            

