import os
import mysql.connector
from datetime import date, datetime
from hijri_converter import Hijri, Gregorian
from PyQt5.QtCore import Qt
from PyQt5 import QtWidgets
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (QApplication,
                             QCheckBox,
                             QComboBox,
                             QDateEdit,
                             QDateTimeEdit,
                             QDial,
                             QDoubleSpinBox,
                             QFontComboBox,
                             QLabel,
                             QLCDNumber,
                             QLineEdit,
                             QMainWindow,
                             QProgressBar,
                             QPushButton,
                             QRadioButton,
                             QSlider,
                             QSpinBox,
                             QTimeEdit,
                             QVBoxLayout,
                             QWidget,
                             QHBoxLayout,
                             QGridLayout,
                             QTableWidget,
                             QHeaderView, QMenu, QAction, QInputDialog, QFormLayout, QDialogButtonBox, QDialog
                             )
from PyQt5.QtGui import (QPalette,
                         QColor,
                         QFont,
                         )

import sys
from random import choice

import utils
from utils import errPopup

mydb = utils.connection
mycursor = mydb.cursor()






sqlFormula = "INSERT INTO Customer (name,phnNo,email,company,GSTIN,address,oldCash) VALUES (%s,%s,%s,%s,%s,%s,%s) "



class Box(QWidget):

    def __init__(self, color):
        super(Box, self).__init__()
        self.setAutoFillBackground(True)

        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(color))
        self.setPalette(palette)


class addClient(QMainWindow):



    def __init__(self):
        super().__init__()
        self.setWindowTitle("Add Client")
        self.setGeometry(150, 150, 500, 300)

        form = QFormLayout()

        self.accNo = self.accNum()

        self.label = QLabel(self.accNo)
        
        self.accNoT = QLabel("Acc No")
        self.nameT = QLabel("Name")
        self.phnT = QLabel("Phone No")
        self.emailT = QLabel("Email")
        self.companyT = QLabel("Company")
        self.vatNoT = QLabel("GST No")
        self.addressT = QLabel("Address")
        self.cashBalT = QLabel("Old Cash Balance")
        # self.goldBalT = QLabel("Old Gold Balance")


        form.addRow(self.accNoT, self.label)
        self.name = QLineEdit()
        form.addRow(self.nameT, self.name)
        self.phnNo = QLineEdit()
        form.addRow(self.phnT, self.phnNo)
        self.email = QLineEdit()
        form.addRow(self.emailT, self.email)
        self.company = QLineEdit()
        form.addRow(self.companyT, self.company)
        self.vatNo = QLineEdit()
        form.addRow(self.vatNoT, self.vatNo)
        self.address = QLineEdit()
        form.addRow(self.addressT, self.address)
        self.oldCash = QLineEdit()
        form.addRow(self.cashBalT, self.oldCash)
        # self.oldGold = QLineEdit()
        # form.addRow(self.goldBalT, self.oldGold)

        self.saveBtn = QPushButton("Save")
        self.saveBtn.clicked.connect(self.pushed)


        screen = QVBoxLayout()
        screen.addLayout(form)
        screen.addWidget(self.saveBtn)

        widget = QWidget()
        widget.setLayout(screen)
        self.setCentralWidget(widget)
        
    def toggle_translate(self):
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
            
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            
            
            self.accNoT.setText(ws["B3"].value)
            self.nameT.setText(ws["B4"].value)
            self.phnT.setText(ws["B23"].value)
            self.emailT.setText(ws["B59"].value)
            self.companyT.setText(ws["B72"].value)
            # self.vatNoT.setText(ws["B64"].value)
            self.phnT.setText(ws["B74"].value)
            self.emailT.setText(ws["B75"].value)
            self.companyT.setText(ws["B76"].value)

            
            # self.popup.label.setText("أدخل كلمة المرور للتغيير")
            self.saveBtn.setText(ws["B7"].value)
            # self.cancelBtn.setText(ws["B73"].value)


    def accNum(self):

        queryLast = "SELECT accNo FROM Customer ORDER BY accNO DESC LIMIT 1"
        mycursor.execute(queryLast)
        try:
            accNo = str(mycursor.fetchone()[0] + 1)
        except TypeError:
            accNo = "1"

        return accNo

    def pushed(self):
        try:
            if(not self.name.text() or self.name.text() == None):
                self.dialog = errPopup("Name field cannot be empty")
                self.dialog.show()

            else:


                self.dialog = popup(self.label.text(), self.name.text(), self.phnNo.text(),
                                    self.email.text(), self.company.text(), self.vatNo.text(),
                                    self.address.text(),
                                    float(self.oldCash.text()), self.label)
                self.dialog.show()

                self.name.clear()
                self.phnNo.clear()
                self.email.clear()
                self.company.clear()
                self.address.clear()
                # self.oldGold.clear()
                self.oldCash.clear()
                self.vatNo.clear()
        except ValueError:
            self.dialog = errPopup("Please put a number in the balance field")
            self.dialog.show()



class popup(QWidget):
    def __init__(self,accNo,name,phnNo,email,company, vatNo, address,oldCash, label):
        super().__init__()
        # self.oldGold = str(oldGold)
        self.oldCash = str(oldCash)
        self.address = address
        self.vatNo = vatNo
        self.company = company
        self.email = email
        self.phnNo = phnNo
        self.name = name
        self.accNo = accNo
        self.Label = label
        layout = QVBoxLayout()
        self.label = QLabel("Please check details.")

        # create a method that the form clears up after saving the details
        self.accNoT = QLabel("Acc No")
        self.nameT = QLabel("Name")
        self.phnT = QLabel("Phone No")
        self.emailT = QLabel("Email")
        self.companyT = QLabel("Company")
        self.vatNoT = QLabel("GST No")
        self.addressT = QLabel("Address")
        self.cashBalT = QLabel("Old Cash Balance")
        # self.goldBalT = QLabel("Old Gold Balance")
        
        form = QFormLayout()
        form.addRow(self.accNoT, QLabel(self.accNo))
        form.addRow(self.nameT, QLabel(name))
        form.addRow(self.phnT, QLabel(phnNo))
        form.addRow(self.emailT, QLabel(email))
        form.addRow(self.companyT, QLabel(company))
        form.addRow(self.vatNoT, QLabel(vatNo))
        form.addRow(self.addressT, QLabel(address))
        form.addRow(self.cashBalT, QLabel(self.oldCash))
        # form.addRow(self.goldBalT, QLabel(self.oldGold))

        self.okBtn = QPushButton("Ok")
        self.cancelBtn = QPushButton("Cancel")

        self.okBtn.clicked.connect(self.pushed)
        self.cancelBtn.clicked.connect(self.wrapper2)

        form.addRow(self.okBtn, self.cancelBtn)

        layout.addWidget(self.label)
        layout.addLayout(form)
        self.setLayout(layout)
        self.toggle_translate()
        
    def toggle_translate(self):
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
            
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            
            
            self.accNoT.setText(ws["B3"].value)
            self.nameT.setText(ws["B4"].value)
            self.phnT.setText(ws["B23"].value)
            self.emailT.setText(ws["B59"].value)
            self.companyT.setText(ws["B72"].value)
            # self.vatNoT.setText(ws["B64"].value)
            self.address.setText(ws["B74"].value)
            self.emailT.setText(ws["B75"].value)
            self.companyT.setText(ws["B76"].value)

            
            self.label.setText("يرجى التحقق من التفاصيل.")
            self.okBtn.setText(ws["B67"].value)
            self.cancelBtn.setText(ws["B73"].value)

    def wrapper1(self):
        self.pushed()

    def wrapper2(self):
        self.close()

    def pushed(self):

        try:
            details = []
            details.append(self.name)
            details.append(self.phnNo)
            details.append(self.email)
            details.append(self.company)
            details.append(self.vatNo)
            details.append(self.address)
            details.append(self.oldCash)
            # details.append(self.oldGold)

            mycursor.execute(sqlFormula, details)
            mydb.commit()
        except TypeError:
            self.dialog = errPopup("16")
            self.dialog.show()
        details_2 = []
        details_2.append(str(date.today()))
        details_2.append(self.accNo)
        if float(self.oldCash) >=0:
            details_2.append(0.00 + float(self.oldCash))
        else:
            details_2.append(0.00)

        if float(self.oldCash) < 0:
            details_2.append((0.00 + float(self.oldCash)) * (-1))
        else:
            details_2.append(0.00)


        details_2.append("Previous Balance")
        details_2.append(-1)

        sqlFormula1 = "INSERT INTO Cust_Bal (date, accNo, cashDebit,cashCredit,billType, billNo) VALUES (%s,%s,%s,%s,%s,%s) "

        try:
            mycursor.execute(sqlFormula1, details_2)
            mydb.commit()
            label = int(self.Label.text()) + 1
            self.Label.setText(str(label))
        except TypeError:
            self.dialog = errPopup("17")
            self.dialog.show()



        self.close()




