import asyncio
import locale
import os
import subprocess

# import win32api

import tempfile
import threading
from datetime import date, datetime

import PyQt5
# import ghostscript
import mysql.connector
from PyQt5 import QtWidgets
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch, cm
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from hijri_converter import Hijri, Gregorian
from num2words import num2words
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
import sys
from PyQt5.QtWidgets import (QApplication,
                             QCheckBox,
                             QComboBox,
                             QDateEdit,
                             QDateTimeEdit,
                             QDial,
                             QDoubleSpinBox,
                             QFontComboBox,
                             QLabel,
                             QLCDNumber,
                             QLineEdit,
                             QMainWindow,
                             QProgressBar,
                             QPushButton,
                             QRadioButton,
                             QSlider,
                             QSpinBox,
                             QTimeEdit,
                             QVBoxLayout,
                             QWidget,
                             QHBoxLayout,
                             QGridLayout,
                             QTableWidget,
                             QHeaderView, QMenu, QAction, QInputDialog, QTableWidgetItem, QFormLayout
                             )
from PyQt5.QtGui import (QPalette,
                         QColor,
                         QFont,
                         )

import sys
from random import choice

import utils






class Box(QWidget):

    def __init__(self, color):
        super(Box, self).__init__()
        self.setAutoFillBackground(True)

        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(color))
        self.setPalette(palette)


class MyTable(QTableWidget):
    def __init__(self, r, c):
        super().__init__(r, c)
        self.init_ui()

    def init_ui(self):
        self.show()

class totalBalance(QMainWindow):
    def __init__(self):
        global mydb
        global mycursor

        mydb = utils.connection
        mycursor = mydb.cursor()
        super().__init__()

        self.setWindowTitle("Customer Balance")

        self.setGeometry(150, 150, 1400, 600)

        header = QHBoxLayout()
        self.header_label = QLabel("Customer Balance")
        self.header_label.setFixedSize(400, 30)
        self.header_label.setFont(QFont('Arial', 20))

        self.header_label.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        header.addWidget(self.header_label)

        iconBox = QGridLayout()
        iconBox_wid1 = Box("red")
        button_row1 = QHBoxLayout(iconBox_wid1)
        self.yearCombo = QComboBox()
        for i in range(2024, 2050):
            self.yearCombo.addItem(str(i))
            self.yearCombo.currentIndexChanged.connect(lambda: self.updateBalance())

        self.pdfBtn = QPushButton("PDF")
        self.pdfBtn.clicked.connect(self.pdf)

        self.printBtn = QPushButton("Print")
        self.printBtn.clicked.connect(self.printdef)

        button_row1.addWidget(self.yearCombo)
        button_row1.addWidget(self.pdfBtn)
        button_row1.addWidget(self.printBtn)
        iconBox_wid1.setFixedHeight(50)
        iconBox.addWidget(iconBox_wid1, 0, 0)

        sheet_box = Box("White")
        sheetV = QVBoxLayout(sheet_box)
        rowQuery = "SELECT COUNT(name) from Customer"
        mycursor.execute(rowQuery)
        res = mycursor.fetchone()
        total_rows = res[0]
        sheet_box.setMaximumHeight(800)
        self.sheet = MyTable(total_rows, 4)
        col_headers = ["Customer Name", "Cash Debit", "Cash Credit", "Cash Balance"]
        sheetV.setSpacing(0)
        sheetV.setContentsMargins(0, 0, 0, 0)
        self.sheet.setHorizontalHeaderLabels(col_headers)
        horiz_header = self.sheet.horizontalHeader()
        horiz_header.setSectionResizeMode(0, QHeaderView.Stretch)
        self.sheet.setColumnWidth(1, 250)
        self.sheet.setColumnWidth(2, 250)
        self.sheet.setColumnWidth(3, 250)
        



        mycursor.execute("SELECT name FROM Customer")
        expName = mycursor.fetchall()
        for i in range(total_rows):
            item = QTableWidgetItem(expName[i][0])
            self.sheet.setItem(i, 0, item)
        year = self.yearCombo.currentText()
        
        totalCashBal = 0

        totalCashDeb = 0
        totalCashCred = 0
 

        for i in range(total_rows):
            query = "SELECT SUM(cashDebit), SUM(cashCredit) FROM  Cust_Bal WHERE accNo = '" + str(i + 1) + "' AND Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31" + "' ORDER BY date ASC"
            mycursor.execute(query)
            sum = mycursor.fetchall()

            print(sum)

            try:
                cDebit = round(float(sum[0][0]), 2)
                totalCashDeb += cDebit

            except(TypeError):
                cDebit = 0.0
                totalCashDeb += cDebit

            try:
                cCredit = round(float(sum[0][1]), 2)
                totalCashCred += cCredit

            except(TypeError):
                cCredit = 0.0
                totalCashCred += cCredit


            try:
                cBal = round(cDebit - cCredit, 2)
            except(TypeError):
                cBal = 0

            try:
                totalCashBal += cBal
            except(TypeError):
                totalCashBal += 0


            cBal = round(cBal,2)

            
            item1 = QTableWidgetItem(str(cDebit))
            item2 = QTableWidgetItem(str(cCredit))
            item3 = QTableWidgetItem(str(cBal))
            self.sheet.setItem(i, 1, item1)
            self.sheet.setItem(i, 2, item2)
            self.sheet.setItem(i, 3, item3)
            


        # query = "SELECT SUM(Amount) FROM expensebalance WHERE  Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31'"
        # mycursor.execute(query)
        # sum = mycursor.fetchall()
        # try:
        #     expense = float(sum[0][0])
        # except:
        #     expense = 0
        
        # self.sheet.setItem(total_rows, 0, QTableWidgetItem("Expenses"))
        # self.sheet.setItem(total_rows, 1, QTableWidgetItem(str(0)))
        # self.sheet.setItem(total_rows, 2, QTableWidgetItem(str(expense)))
        # self.sheet.setItem(total_rows, 3, QTableWidgetItem(str(expense * (-1))))
        # totalCashCred += expense
        # totalCashBal += expense * (-1)
        
        print(totalCashBal)
        

        self.sheet.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)

        sheetV.addWidget(self.sheet)

        total_sheet_box = Box("White")
        total_sheetV = QVBoxLayout(total_sheet_box)

        total_sheet_box.setMaximumHeight(65)
        self.total_sheet = MyTable(1, 4)

        total_sheetV.setSpacing(0)
        total_sheetV.setContentsMargins(0, 0, 0, 0)

        horiz_header = self.total_sheet.horizontalHeader()
        self.total_sheet.horizontalHeader().setVisible(False)
        self.total_sheet.verticalHeader().setVisible(False)
        horiz_header.setSectionResizeMode(0, QHeaderView.Stretch)
        self.total_sheet.setColumnWidth(1, 250)
        self.total_sheet.setColumnWidth(2, 250)
        self.total_sheet.setColumnWidth(3, 250)
        
        itemm = QTableWidgetItem("Total")
        self.total_sheet.setItem(0,0,itemm)
        # itemm1 = QTableWidgetItem("Balance")
        # self.total_sheet.setItem(1, 0, itemm1)
        
        self.toggle_translate()

        totalCashBal = round(totalCashBal, 2)
        # totalGoldBal = round(totalGoldBal, 3)
        totalCashDeb = round(totalCashDeb,2)
        totalCashCred = round(totalCashCred,2)
        # totalGoldDeb = round(totalGoldDeb, 2)
        # totalGoldCred = round(totalGoldCred,2)
        

        itemm1 = QTableWidgetItem(str(totalCashDeb))
        itemm2 = QTableWidgetItem(str(totalCashCred))
        itemm3 = QTableWidgetItem(str(totalCashBal))
        self.total_sheet.setItem(0,1,itemm1)
        self.total_sheet.setItem(0,2,itemm2)
        self.total_sheet.setItem(0,3,itemm3)
        

        # if totalCashBal > 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(totalCashBal))
        #     self.total_sheet.setItem(1, 1, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 2, item1)

        # if totalCashBal < 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(totalCashBal))
        #     self.total_sheet.setItem(1, 2, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 1, item1)

        # if totalCashBal == 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 2, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 1, item1)

        # if totalGoldBal > 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(totalGoldBal))
        #     self.total_sheet.setItem(1, 3, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 4, item1)

        # if totalGoldBal < 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(totalGoldBal))
        #     self.total_sheet.setItem(1, 4, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 3, item1)

        # if totalGoldBal == 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 4, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 3, item1)

        self.total_sheet.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)

        total_sheetV.addWidget(self.total_sheet)

        mainLayout = QVBoxLayout()
        mainLayout.addLayout(header)
        mainLayout.addLayout(iconBox)
        mainLayout.addWidget(sheet_box)
        mainLayout.addWidget(total_sheet_box)

        widget = QWidget()
        widget.setLayout(mainLayout)
        self.setCentralWidget(widget)
        
        

    
    def toggle_translate(self):
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
            
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            
            self.header_label.setText(ws["B33"].value)
            
            self.pdfBtn.setText("PDF")
            self.printBtn.setText(ws["B9"].value)
             
            col_header_arabic = []
            col_header_arabic.append(ws["B34"].value)
            col_header_arabic.append(ws["B37"].value)
            col_header_arabic.append(ws["B38"].value)
            col_header_arabic.append(ws["B35"].value)
            col_header_arabic.append(ws["B36"].value)
            self.sheet.setHorizontalHeaderLabels(col_header_arabic)
            
            item1 = QTableWidgetItem(ws["B15"].value)
            item2 = QTableWidgetItem(ws["B68"].value)
            self.total_sheet.clear()
            self.total_sheet.setItem(0, 0, item1)
            self.total_sheet.setItem(1, 0, item2)
            
            
            
    
       

    def updateBalance(self):
        self.sheet.clear()
        self.total_sheet.clear()

        rowQuery = "SELECT COUNT(Name) from Customer"
        mycursor.execute(rowQuery)
        res = mycursor.fetchone()
        total_rows = res[0]
        col_headers = ["Customer Name", "Cash Debit", "Cash Credit", "Cash Balance"]
        self.sheet.setHorizontalHeaderLabels(col_headers)
        horiz_header = self.sheet.horizontalHeader()
        horiz_header.setSectionResizeMode(0, QHeaderView.Stretch)
        self.sheet.setColumnWidth(1, 250)
        self.sheet.setColumnWidth(2, 250)
        self.sheet.setColumnWidth(3, 250)
        
        mycursor.execute("SELECT name FROM Customer")
        expName = mycursor.fetchall()
        for i in range(total_rows):
            item = QTableWidgetItem(expName[i][0])
            self.sheet.setItem(i, 0, item)
        year = self.yearCombo.currentText()
        totalCashBal = 0
        # totalGoldBal = 0
        totalCashDeb = 0
        totalCashCred = 0
        # totalGoldDeb = 0
        # totalGoldCred = 0

        for i in range(total_rows):
            query = "SELECT SUM(cashDebit), SUM(cashCredit) FROM  Cust_Bal WHERE accNo = '" + str(i + 1) + "' AND Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31" + "' ORDER BY date ASC"
            mycursor.execute(query)
            sum = mycursor.fetchall()

            print(sum)

            try:
                cDebit = round(sum[0][0], 2)
                totalCashDeb += cDebit

            except(TypeError):
                cDebit = 0
                totalCashDeb += cDebit

            try:
                cCredit = round(sum[0][1], 2)
                totalCashCred += cCredit

            except(TypeError):
                cCredit = 0
                totalCashCred += cCredit

            # try:
            #     gDebit = round(sum[0][2], 2)

            # except(TypeError):
            #     gDebit = 0

            # try:
            #     gCredit = round(sum[0][3], 2)

            # except(TypeError):
            #     gCredit = 0

            try:
                cBal = round(cDebit - cCredit, 2)
            except(TypeError):
                cBal = 0
            # try:
            #     gBal = round(gDebit - gCredit, 2)
            # except(TypeError):
            #     gBal = 0
            try:
                totalCashBal += cBal
            except(TypeError):
                totalCashBal += 0
            # try:
            #     totalGoldBal += gBal
            # except(TypeError):
            #     totalGoldBal += 0

            cBal = round(cBal,2)
            # gBal = round(gBal,3)
            
            item1 = QTableWidgetItem(str(cDebit))
            item2 = QTableWidgetItem(str(cCredit))
            item3 = QTableWidgetItem(str(cBal))
            self.sheet.setItem(i, 1, item1)
            self.sheet.setItem(i, 2, item2)
            self.sheet.setItem(i, 3, item3)

        self.sheet.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)



        horiz_header = self.total_sheet.horizontalHeader()
        horiz_header.setSectionResizeMode(0, QHeaderView.Stretch)
        self.total_sheet.setColumnWidth(1, 250)
        self.total_sheet.setColumnWidth(2, 250)
        self.total_sheet.setColumnWidth(3, 250)
        
        self.total_sheet.horizontalHeader().setVisible(False)
        self.total_sheet.verticalHeader().setVisible(False)
        itemm = QTableWidgetItem("Total")
        self.total_sheet.setItem(0, 0, itemm)
        # itemm1 = QTableWidgetItem("Balance")
        # self.total_sheet.setItem(1, 0, itemm1)
        
        self.toggle_translate()

        totalCashBal = round(totalCashBal, 2)
        # totalGoldBal = round(totalGoldBal, 3)
        totalCashDeb = round(totalCashDeb,2)
        totalCashCred = round(totalCashCred,2)
        # totalGoldDeb = round(totalGoldDeb, 2)
        # totalGoldCred = round(totalGoldCred,2)
        

        itemm1 = QTableWidgetItem(str(totalCashDeb))
        itemm2 = QTableWidgetItem(str(totalCashCred))
        itemm3 = QTableWidgetItem(str(totalCashBal))
        self.total_sheet.setItem(0,1,itemm1)
        self.total_sheet.setItem(0,2,itemm2)
        self.total_sheet.setItem(0,3,itemm3)


        # if totalCashBal > 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(totalCashBal))
        #     self.total_sheet.setItem(1, 1, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 2, item1)

        # if totalCashBal < 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(totalCashBal))
        #     self.total_sheet.setItem(1, 2, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 1, item1)

        # if totalCashBal == 0:
        #     item0 = QTableWidgetItem(str(totalCashDeb))
        #     self.total_sheet.setItem(0, 1, item0)
        #     item01 = QTableWidgetItem(str(totalCashCred))
        #     self.total_sheet.setItem(0, 2, item01)
        #     item = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 2, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 1, item1)

        # if totalGoldBal > 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(totalGoldBal))
        #     self.total_sheet.setItem(1, 3, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 4, item1)

        # if totalGoldBal < 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(totalGoldBal))
        #     self.total_sheet.setItem(1, 4, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 3, item1)

        # if totalGoldBal == 0:
        #     item0 = QTableWidgetItem(str(totalGoldDeb))
        #     self.total_sheet.setItem(0, 3, item0)
        #     item01 = QTableWidgetItem(str(totalGoldCred))
        #     self.total_sheet.setItem(0, 4, item01)
        #     item = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 4, item)
        #     item1 = QTableWidgetItem(str(0))
        #     self.total_sheet.setItem(1, 3, item1)

        self.total_sheet.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        

    def printdef(self):
        self.pdf()
        Date = str(date.today())
        year = self.yearCombo.currentText()

        if str(date.today().year) != year:
            Datee = year + "-12-31"
            my_path = 'TotalBalance\\' + year + "-" + Datee + '.pdf'
        else:
            my_path = 'TotalBalance\\' + year + "-" + Date + '.pdf'
        location = os.path.join(os.getcwd(), my_path).replace('\\', '\\\\')
        # import webbrowser
        # new = 2
        # webbrowser.open(location, new=new)
        # os.remove(location)






        GSPRINT_PATH = os.path.join(os.getcwd(), "GSPRINT\\gsprint.exe")

        # # YOU CAN PUT HERE THE NAME OF YOUR SPECIFIC PRINTER INSTEAD OF DEFAULT
        # currentprinter = win32print.GetDefaultPrinter()
        #
        # win32api.ShellExecute(0, 'open', GSPRINT_PATH,
        #                       '-ghostscript "' + GHOSTSCRIPT_PATH + '" -printer "' + currentprinter + '" "' + location + '"',
        #                       '.', 0)
        GHOSTSCRIPT_PATH = os.path.join(os.getcwd(), "gs\\gs10.01.2\\bin\\gswin64c")
        if sys.platform == 'win32':
            args = f'"{GHOSTSCRIPT_PATH}" ' \
                   '-sDEVICE=mswinpr2 ' \
                   '-dBATCH ' \
                   '-dNOPAUSE ' \
                   '-dFitPage ' \
                   '-dNOPROMPT ' \
                   '-dPrinted ' \
                   '-sOutputFile="%printer%myPrinterName" '
            ghostscript = args + os.path.join(os.getcwd(), location).replace('\\', '\\\\')
            subprocess.call(ghostscript, shell=True)

            os.remove(location)

        # args = [
        #     "-dPrinted", "-dBATCH", "-dNOSAFER", "-dNOPAUSE", "-dNOPROMPT", "-dPDFFitPage",
        #
        #     "-dNumCopies=1",
        #     "-sDEVICE=mswinpr2",
        #     f'-sOutputFile="%printer%{win32print.GetDefaultPrinter()}"',
        #     f'"{f}"'
        # ]
        #
        # encoding = locale.getpreferredencoding()
        # args = [a.encode(encoding) for a in args]
        # ghostscript.Ghostscript(*args)


    def pdf(self):
        Date = str(date.today())
        year = self.yearCombo.currentText()
        dayOne = "01/01/" + year

        if str(date.today().year) != year:
            Datee = year + "-12-31"
            my_path = 'TotalBalance/' + year + "-" + Datee + '.pdf'
        else:
            my_path = 'TotalBalance/' + year + "-" + Date + '.pdf'
        c = canvas.Canvas(my_path, bottomup=1, pagesize=A4)
        c.setStrokeColorRGB(0, 0, 0)
        c.setFont('Helvetica', 25)
        c.drawString(230, 800, "Total Balance")
        c.setFont('Helvetica', 12)
        c.drawString(240, 770, dayOne + " - " + Date)

        # prerequisites
        rowQuery = "SELECT COUNT(*) from Customer"
        mycursor.execute(rowQuery)
        res = mycursor.fetchone()
        total_rows = res[0]
        print(total_rows)
        mycursor.execute("SELECT * FROM Customer")
        expName = mycursor.fetchall()
        totalCashBal = 0
        totalGoldBal = 0
        
        pdfmetrics.registerFont(TTFont('Rupee', 'ITF-Rupee.ttf'))
        c.setFont('Helvetica', 9)
        for i in range(total_rows):

            if (i == 31):
                reqTimes = total_rows - i
                self.addNewPage(c, reqTimes, 31)
                break
            
            # headings
            
            # mainbox
            c.rect(.40 * cm, 24.30 * cm + (.8 * cm), 20.1 * cm, .8 * cm, fill=0)

            # seperator1
            c.line(150 + 40, 689 + (22.65), 150 + 40, 712 + (22.65))

            # seperator2
            c.line(250 + 40 + 33, 689 + (22.65), 250 + 40 + 33, 712 + (22.65))

            # seperator3
            c.line(350 + 40 + 66, 689 + (22.65), 350 + 40 + 66, 712 + (22.65))


            # ID
            c.drawString(15, 696 + (22.65), "Id")
            # seperator1
            c.line(40, 689 + (22.65), 40, 712 + (22.65))
            # seperator1
            c.line(40, 689 - (22.65 * i), 40, 712 - (22.65 * i))
            # Customer Name
            c.drawString(15, 696 - (22.65 * i), str(expName[i][0]))


            # Customer Name
            c.drawString(15 + 40, 696 + (22.65), "Customer Name")
            c.drawString(155 + 40, 696 + (22.65), "Cash Debit")
            c.drawString(255 + 40 + 33, 696 + (22.65), "Cash Credit")
            c.drawString(355 + 40 + 66, 696 + (22.65), "Cash Balance")

            
            # sheet body
            # mainbox
            c.rect(.40 * cm, 24.30 * cm - (.8 * cm * i), 20.1 * cm, .8 * cm, fill=0)

            # seperator1
            c.line(150 + 40, 689 - (22.65 * i), 150 + 40, 712 - (22.65 * i))

            # seperator2
            c.line(250 + 40 + 33, 689 - (22.65 * i), 250 + 40 + 33, 712 - (22.65 * i))

            # seperator3
            c.line(350 + 40 + 66, 689 - (22.65 * i), 350 + 40 + 66, 712 - (22.65 * i))


            # Customer Name
            c.drawString(15 + 40, 696 - (22.65 * i), expName[i][1])

            query = "SELECT SUM(cashDebit), SUM(cashCredit) FROM Cust_Bal WHERE accNo = '" + str(
                i + 1) + "' AND Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31" + "' ORDER BY date ASC"
            mycursor.execute(query)
            sum = mycursor.fetchall()
            try:
                cDebit = round(sum[0][0], 2)

            except(TypeError):
                cDebit = 0

            try:
                cCredit = round(sum[0][1], 2)

            except(TypeError):
                cCredit = 0


            try:
                cBal = round(cDebit - cCredit, 2)
            except(TypeError):
                cBal = 0

            try:
                totalCashBal += cBal
            except(TypeError):
                totalCashBal += 0


            cBal = round(cBal,2)

            
            c.drawString(155 + 50, 696 - (22.65 * i), str(cDebit)) # Debit
            c.drawString(255 + 50 + 33, 696 - (22.65 * i), str(cCredit)) # Credit
            c.drawString(355 + 50 + 66, 696 - (22.65 * i), str(cBal))  # Balance
            c.setFont('Rupee', 10)
            c.drawString(155 + 40, 696 - (22.65 * i), "I") # Debit
            c.drawString(255 + 40 + 33, 696 - (22.65 * i), "I") # Credit
            c.drawString(355 + 40 + 66, 696 - (22.65 * i), "I")  # Balance
            c.setFont('Helvetica', 9)



            # for the total line
            if i == total_rows - 1 and i < 40:
                print(i)
                # for the for loop
                rows = self.total_sheet.rowCount()
                col = self.total_sheet.columnCount()

                rowAr = []
                sheetAr = []

                for a in range(rows):
                    rowAr = []
                    for b in range(col):
                        rowAr.append(self.total_sheet.item(a, b).text())
                    sheetAr.append(rowAr)

                # c.line(40, 689 - (22.65 * (i + 1)), 40, 712 - (22.65 * (i + 1)))
                # c.drawString(15, 696 - (22.65 * (i + 1)), str(0))

                for k in range(1):
                    # mainbox
                    c.rect(.40 * cm, 24.30 * cm - (.8 * cm * (i + 1 + k)), 20.1 * cm, .8 * cm, fill=0)

                    # seperator1
                    c.line(150 + 40, 689 - (22.65 * (i + 1 + k)), 150 + 40, 712 - (22.65 * (i + 1)))

                    # seperator2
                    c.line(250 + 40 + 33, 689 - (22.65 * (i + 1 + k)), 250 + 40 + 33, 712 - (22.65 * (i + 1)))

                    # seperator3
                    c.line(350 + 40 + 66, 689 - (22.65 * (i + 1 + k)), 350 + 40 + 66, 712 - (22.65 * (i + 1)))

                    # seperator4
                    # c.line(450 + 40, 689 - (22.65 * (i + 1 + k)), 450 + 40, 712 - (22.65 * (i + 1)))

                # Total
                # c.drawString(15 + 40, 696 - (22.65 * (i + 1)), "Expenses")
                # Total
                c.drawString(15 + 40, 696 - (22.65 * (i + 1)), "Total")
                
                # query = "SELECT SUM(Amount) FROM expensebalance WHERE  Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31'"
                # mycursor.execute(query)
                # sum = mycursor.fetchall()
                # print(sum)
                # try:
                #     expense = float(sum[0][0])
                # except:
                #     expense = 0
                
                # c.drawString(155 + 40, 696 - (22.65 * (i+1)), "SAR " + str(0))
                # c.drawString(255 + 40 + 33, 696 - (22.65 * (i+1)), "SAR " + str(expense))
                # c.drawString(355 + 40 + 66, 696 - (22.65 * (i+1)), "SAR " + str(expense * (-1)))

                # for loop to put up the values
                for j in range(len(sheetAr)):
                    c.drawString(155 + 50, 696 - (22.65 * (i+1 + j)), str(sheetAr[j][1]))
                    c.drawString(255 + 50 + 33, 696 - (22.65 * (i+1 + j)), str(sheetAr[j][2]))
                    c.drawString(355 + 50 + 66, 696 - (22.65 * (i+1 + j)), str(sheetAr[j][3]))
                    c.setFont('Rupee', 10)
                    c.drawString(155 + 40, 696 - (22.65 * (i+1 + j)), "I") # Debit
                    c.drawString(255 + 40 + 33, 696 - (22.65 * (i+1 + j)), "I") # Credit
                    c.drawString(355 + 40 + 66, 696 - (22.65 * (i+1 + j)), "I")  # Balance
                    c.setFont('Helvetica', 9)
                    








        c.showPage()  # saves current page
        c.save()


    def addNewPage(self, c, reqTimes, timesBeen):
        # this method is to add rows and values to excess pages
        c.showPage()
        c.setPageSize(A4)

        # prerequisites
        year = self.yearCombo.currentText()
        rowQuery = "SELECT COUNT(*) from Customer"
        mycursor.execute(rowQuery)
        res = mycursor.fetchone()
        total_rows = res[0]
        mycursor.execute("SELECT * FROM Customer")
        expName = mycursor.fetchall()
        totalCashBal = 0
        totalGoldBal = 0


        for i in range(reqTimes):

            if (i == 35):
                self.addNewPage(c, (reqTimes - i), (timesBeen + i))
                break

            # mainbox
            c.rect(.40 * cm, 28.30 * cm + (.8 * cm), 20.1 * cm, .8 * cm, fill=0)

            # seperator1
            c.line(15 + 40, 802 + (22.65), 150 + 40, 825 + (22.65))

            # seperator2
            c.line(250 + 40 + 33, 802 + (22.65), 250 + 40 + 33, 825 + (22.65))

            # seperator3
            c.line(350 + 40 + 66, 802 + (22.65), 350 + 40 + 66, 825 + (22.65))

     

            # ID
            c.drawString(15, 809 + (22.65), "Id")
            # seperator1
            c.line(40, 802 + (22.65), 40, 825 + (22.65))
            # seperator1
            c.line(40, 802 - (22.65 * i), 40, 825 - (22.65 * i))
            # Customer Name
            c.drawString(15, 809 - (22.65 * i), str(expName[i][0]))

            # Customer Name
            c.drawString(15 + 40, 809 + (22.65), "Customer Name")
            c.drawString(155 + 40, 809 + (22.65), "Cash Debit")
            c.drawString(255 + 40 + 33, 809 + (22.65), "Cash Credit")
            c.drawString(355 + 40 + 66, 809 + (22.65), "Cash Balance")
         

            # mainbox
            c.rect(.40 * cm, 28.30 * cm - (.8 * cm * i), 20.1 * cm, .8 * cm, fill=0)

            # seperator1
            c.line(150 + 40, 825 - (22.65 * i), 150 + 40, 802 - (22.65 * i))

            # seperator2
            c.line(250 + 40 + 33, 825 - (22.65 * i), 250 + 40 + 33, 802 - (22.65 * i))

            # seperator3
            c.line(350 + 40 + 66, 825 - (22.65 * i), 350 + 40 + 66, 802 - (22.65 * i))


            # Customer Name
            c.drawString(15 + 40, 809 - (22.65 * i), expName[i][1])

            query = "SELECT SUM(cashDebit), SUM(cashCredit), SUM(goldDebit), SUM(goldCredit) FROM Cust_Bal WHERE accNo = '" + str(
                i + 1) + "' AND Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31" + "'  ORDER BY date ASC"
            mycursor.execute(query)
            sum = mycursor.fetchall()
            try:
                cDebit = round(sum[0][0], 2)

            except(TypeError):
                cDebit = 0

            try:
                cCredit = round(sum[0][1], 2)

            except(TypeError):
                cCredit = 0



            try:
                cBal = round(cDebit - cCredit, 2)
            except(TypeError):
                cBal = 0

            try:
                totalCashBal += cBal
            except(TypeError):
                totalCashBal += 0

            cBal = round(cBal,2)

            c.drawString(155 + 50, 809 - (22.65 * i), str(cDebit))
            c.drawString(255 + 50 + 33, 809 - (22.65 * i), str(cCredit))
            c.drawString(355 + 50 + 66, 809 - (22.65 * i), str(cBal))
            c.setFont('Rupee', 10)
            c.drawString(155 + 50, 809 - (22.65 * i), "I") # Debit
            c.drawString(255 + 50 + 33, 809 - (22.65 * i), "I") # Credit
            c.drawString(355 + 50 + 66, 809 - (22.65 * i), "I")  # Balance
            c.setFont('Helvetica', 9)


            # for the total line
            if reqTimes - i - 1 == 0 and i < 35:
                if i != 33:
                    print(i)
                    # for the for loop
                    rows = self.total_sheet.rowCount()
                    col = self.total_sheet.columnCount()

                    rowAr = []
                    sheetAr = []

                    for a in range(rows):
                        rowAr = []
                        for b in range(col):
                            rowAr.append(self.total_sheet.item(a, b).text())
                        sheetAr.append(rowAr)
                    
                    # c.line(40, 802 + (22.65 * (i + 1)), 40, 825 + (22.65 * (i + 1)))
                    # c.drawString(15, 809 - (22.65 * (i + 1)), str(0))
                    
                    for k in range(1):
                        # mainbox
                        c.rect(.40 * cm, 28.30 * cm - (.8 * cm * (i + 1 + k)), 20.1 * cm, .8 * cm, fill=0)

                        # seperator1
                        c.line(150 + 40, 825 - (22.65 * (i + 1 + k)), 150 + 40, 802 - (22.65 * (i + 1)))

                        # seperator2
                        c.line(250 + 40 + 33, 825 - (22.65 * (i + 1 + k)), 250 + 40 + 33, 802 - (22.65 * (i + 1)))

                        # seperator3
                        c.line(350 + 40 + 66, 825 - (22.65 * (i + 1 + k)), 350 + 40 + 66, 802 - (22.65 * (i + 1)))


                    # Total
                    # c.drawString(15 + 40, 809 - (22.65 * (i + 1)), "Expenses")
                    # Total
                    c.drawString(15 + 40, 809 - (22.65 * (i + 1)), "Total")

                    # query = "SELECT SUM(Amount) FROM expensebalance WHERE  Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31'"
                    # mycursor.execute(query)
                    # sum = mycursor.fetchall()
                    # print(sum)
                    # try:
                    #     expense = float(sum[0][0])
                    # except:
                    #     expense = 0
                    
                    # c.drawString(155 + 40, 809 - (22.65 * (i + 1)), "SAR " + str(0))
                    # c.drawString(255 + 40 +33, 809 - (22.65 * (i + 1)), "SAR " + str(expense))
                    # c.drawString(355 + 40 + 66, 809 - (22.65 * (i + 1)), "SAR " + str(expense * (-1)))
                    
                    # for loop to put up the values
                    for j in range(len(sheetAr)):
                        c.drawString(155 + 50, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][1]))
                        c.drawString(255 + 50 + 33, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][2]))
                        c.drawString(355 + 50 + 66, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][3]))
                        c.setFont('Rupee', 10)
                        c.drawString(155 + 40, 809 - (22.65 * (i + 1 + j)), "I") # Debit
                        c.drawString(255 + 40 + 33, 809 - (22.65 * (i + 1 + j)), "I") # Credit
                        c.drawString(355 + 40 + 66, 809 - (22.65 * (i + 1 + j)), "I")  # Balance
                        c.setFont('Helvetica', 9)
                        


                else:

                    c.showPage()
                    c.setPageSize(A4)
                    i = 0

                    print(i)
                    # for the for loop
                    rows = self.total_sheet.rowCount()
                    col = self.total_sheet.columnCount()

                    rowAr = []
                    sheetAr = []

                    for a in range(rows):
                        rowAr = []
                        for b in range(col):
                            rowAr.append(self.total_sheet.item(a, b).text())
                        sheetAr.append(rowAr)
                        
                    # c.line(40, 802 + (22.65 * (i + 1)), 40, 825 + (22.65 * (i + 1)))
                    # c.drawString(15, 809 - (22.65 * (i + 1)), str(0))
                    
                    for k in range(1):
                        # mainbox
                        c.rect(.40 * cm, 28.30 * cm - (.8 * cm * (i + 1 + k)), 20.1 * cm, .8 * cm, fill=0)

                        # seperator1
                        c.line(150 + 40, 825 - (22.65 * (i + 1 + k)), 150 + 40, 802 - (22.65 * (i + 1)))

                        # seperator2
                        c.line(250 + 40 + 33, 825 - (22.65 * (i + 1 + k)), 250 + 40 + 33, 802 - (22.65 * (i + 1)))

                        # seperator3
                        c.line(350 + 40 + 66, 825 - (22.65 * (i + 1 + k)), 350 + 40 + 66, 802 - (22.65 * (i + 1)))

                    # query = "SELECT SUM(Amount) FROM expensebalance WHERE  Date BETWEEN " + "'" + "2021-01-01" + "' AND " + "'" + str(year) + "/12/31'"
                    # mycursor.execute(query)
                    # sum = mycursor.fetchall()
                    # print(sum)
                    # try:
                    #     expense = float(sum[0][0])
                    # except:
                    #     expense = 0
                    
                    # c.drawString(155 + 40, 809 - (22.65 * (i + 1)), "SAR " + str(0))
                    # c.drawString(255 + 40 +33, 809 - (22.65 * (i + 1)), "SAR " + str(expense))
                    # c.drawString(355 + 40 + 66, 809 - (22.65 * (i + 1)), "SAR " + str(expense * (-1)))

                    # Total
                    # c.drawString(15 + 40, 809 - (22.65 * (i + 1)), "Expenses")

                    # Total
                    c.drawString(15 + 40, 809 - (22.65 * (i + 1)), "Total")

                    # for loop to put up the values
                    for j in range(len(sheetAr)):
                        c.drawString(155 + 40, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][1]))
                        c.drawString(255 + 40 + 33, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][2]))
                        c.drawString(355 + 40 + 66, 809 - (22.65 * (i + 1 + j)), str(sheetAr[j][3]))
                        c.setFont('Rupee', 10)
                        c.drawString(155 + 40, 809 - (22.65 * (i + 1 + j)), "I") # Debit
                        c.drawString(255 + 40 + 33, 809 - (22.65 * (i + 1 + j)), "I") # Credit
                        c.drawString(355 + 40 + 66, 809 - (22.65 * (i + 1 + j)), "I")  # Balance
                        c.setFont('Helvetica', 9)
                        
