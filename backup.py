import sys
import os
import pickle
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.auth.exceptions import RefreshError  # Import the RefreshError
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import openpyxl

# If modifying these SCOPES, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive.file']

class BackupApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.creds = None
        self.authenticate_google_drive()

    def initUI(self):
        self.setWindowTitle('Data Backup to Google Drive')
        layout = QVBoxLayout()

        self.label = QLabel('Enter password:')
        layout.addWidget(self.label)

        self.data_entry = QLineEdit()
        layout.addWidget(self.data_entry)

        self.submit_button = QPushButton('Submit and Backup')
        self.submit_button.clicked.connect(self.backup_data)
        layout.addWidget(self.submit_button)
        self.setLayout(layout)
        workbook = openpyxl.load_workbook('settings.xlsx')
        worksheet = workbook.active
        if (worksheet["B2"].value == "arabic"):
        
            workbook1 = openpyxl.load_workbook('translation.xlsx')
            ws = workbook1.active
            self.label.setText("أدخل كلمة المرور:")
            self.submit_button.setText("إرسال والنسخ الاحتياطي")
        else:
            pass

    def authenticate_google_drive(self):
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                self.creds = pickle.load(token)
        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                try:
                    self.creds.refresh(Request())
                except RefreshError:  # Use the imported RefreshError
                    print("Token refresh failed. Re-authenticating...")
                    os.remove('token.pickle')
                    flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                    self.creds = flow.run_local_server(port=0)
                    with open('token.pickle', 'wb') as token:
                        pickle.dump(self.creds, token)
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                self.creds = flow.run_local_server(port=0)
                with open('token.pickle', 'wb') as token:
                    pickle.dump(self.creds, token)

    def backup_data(self):
        data = self.data_entry.text()
        if not data:
            workbook = openpyxl.load_workbook('settings.xlsx')
            worksheet = workbook.active
            if (worksheet["B2"].value == "arabic"):
            
                workbook1 = openpyxl.load_workbook('translation.xlsx')
                ws = workbook1.active
                QMessageBox.warning(self, 'Input Error', 'الرجاء إدخال كلمة المرور')
                return
            else:
                QMessageBox.warning(self, 'Input Error', 'Please enter password.')
                return
        elif data == "13/12":
            
            filename = "oilBiller.xlsx"

            try:
                service = build('drive', 'v3', credentials=self.creds)
                # Search for the existing file by name
                results = service.files().list(q=f"name='{filename}' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
                                            spaces='drive',
                                            fields="files(id, name)").execute()
                items = results.get('files', [])

                # If file exists, update it; otherwise, create a new one
                if items:
                    file_id = items[0]['id']
                    media = MediaFileUpload(filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    file = service.files().update(fileId=file_id, media_body=media).execute()
                    QMessageBox.information(self, 'Success', 'Data updated successfully on Google Drive.')
                else:
                    file_metadata = {'name': filename}
                    media = MediaFileUpload(filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                    QMessageBox.information(self, 'Success', 'Data backed up successfully to Google Drive.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'An error occurred: {e}')

        else:
            workbook = openpyxl.load_workbook('settings.xlsx')
            worksheet = workbook.active
            if (worksheet["B2"].value == "arabic"):
            
                workbook1 = openpyxl.load_workbook('translation.xlsx')
                ws = workbook1.active
                QMessageBox.warning(self, 'Input Error', 'الرجاء إدخال كلمة المرور الصحيحة')
                return
            else:
                QMessageBox.warning(self, 'Input Error', 'Please enter correct password.')
                return
            
            
            return
            
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = BackupApp()
    ex.show()
    sys.exit(app.exec_())
